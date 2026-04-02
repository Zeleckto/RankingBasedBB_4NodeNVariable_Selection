"""
results_generator.py — Universal results analysis for RankingBasedBB experiments.

Run AFTER evaluate.py has completed. Reads solve results from pickle files
(or you can paste numbers directly into MANUAL_RESULTS below),
generates all figures, Excel workbook, and a findings summary txt.

Usage:
    python results_generator.py                    # uses saved result pickles
    python results_generator.py --manual           # uses MANUAL_RESULTS dict below
    python results_generator.py --problem setcover
    python results_generator.py --all              # all four benchmarks
"""

import os
import sys
import argparse
import pickle
import numpy as np
import json
from scipy import stats
from datetime import datetime

# ── Optional imports (install if missing) ─────────────────────────────────────
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    HAS_MPL = True
except ImportError:
    HAS_MPL = False
    print("Warning: matplotlib not installed. Skipping figures. pip install matplotlib")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    print("Warning: openpyxl not installed. Skipping Excel. pip install openpyxl")

# ══════════════════════════════════════════════════════════════════════════════
# MANUAL RESULTS — fill these in if not using saved pickle files
# Format: { policy_name: { difficulty: { 'times': [...], 'nodes': [...], 'solved': int, 'total': int } } }
# ══════════════════════════════════════════════════════════════════════════════
MANUAL_RESULTS = {
    # Example structure — replace with your actual numbers
    "setcover": {
        "SCIP Default": {
            "easy":   {"times": [], "nodes": [], "solved": 0, "total": 0},
            "medium": {"times": [], "nodes": [], "solved": 0, "total": 0},
            "hard":   {"times": [], "nodes": [], "solved": 0, "total": 0},
        },
        "GCN + hybridestim": {
            "easy":   {"times": [], "nodes": [], "solved": 0, "total": 0},
            "medium": {"times": [], "nodes": [], "solved": 0, "total": 0},
            "hard":   {"times": [], "nodes": [], "solved": 0, "total": 0},
        },
        "GCN + SCIP UCT": {
            "easy":   {"times": [], "nodes": [], "solved": 0, "total": 0},
            "medium": {"times": [], "nodes": [], "solved": 0, "total": 0},
            "hard":   {"times": [], "nodes": [], "solved": 0, "total": 0},
        },
        "GCN + Neural UCT": {
            "easy":   {"times": [], "nodes": [], "solved": 0, "total": 0},
            "medium": {"times": [], "nodes": [], "solved": 0, "total": 0},
            "hard":   {"times": [], "nodes": [], "solved": 0, "total": 0},
        },
    }
}

POLICY_COLORS = {
    "SCIP Default":      "#888888",
    "GCN + hybridestim": "#2196F3",
    "GCN + SCIP UCT":    "#FF9800",
    "GCN + Neural UCT":  "#4CAF50",
}

POLICY_ORDER = ["SCIP Default", "GCN + hybridestim", "GCN + SCIP UCT", "GCN + Neural UCT"]
DIFFICULTIES  = ["easy", "medium", "hard"]
BENCHMARKS    = ["setcover", "auction", "facility", "indset"]
BENCHMARK_NAMES = {
    "setcover": "Set Covering",
    "auction":  "Combinatorial Auction",
    "facility": "Capacitated Facility Location",
    "indset":   "Maximum Independent Set",
}

# ── Metrics helpers ────────────────────────────────────────────────────────────

def sgm(values, shift=1.0):
    """1-shifted geometric mean."""
    v = np.array([x for x in values if x is not None and not np.isinf(x)], dtype=float)
    if len(v) == 0:
        return float('nan')
    return np.exp(np.mean(np.log(v + shift))) - shift

def win_rate(baseline_times, proposed_times):
    b = np.array(baseline_times)
    p = np.array(proposed_times)
    n = min(len(b), len(p))
    return (p[:n] < b[:n]).mean()

def wilcoxon_test(baseline_times, proposed_times):
    b = np.array(baseline_times)
    p = np.array(proposed_times)
    n = min(len(b), len(p))
    b, p = b[:n], p[:n]
    if len(b) < 3 or np.allclose(b, p):
        return 1.0
    try:
        _, pval = stats.wilcoxon(b, p, alternative='greater')
        return pval
    except Exception:
        return 1.0

def improvement_pct(baseline, proposed):
    if baseline == 0:
        return 0.0
    return (baseline - proposed) / baseline * 100

# ── Load results ───────────────────────────────────────────────────────────────

def load_results(results_dir="results"):
    """Load pickled result dicts saved by evaluate.py."""
    data = {}
    if not os.path.exists(results_dir):
        return data
    for f in os.listdir(results_dir):
        if f.endswith('.pkl'):
            with open(os.path.join(results_dir, f), 'rb') as fp:
                d = pickle.load(fp)
                key = f.replace('.pkl', '')
                data[key] = d
    return data

# ── Figure generation ──────────────────────────────────────────────────────────

def fig_solve_time_comparison(results, problem, outdir):
    """Bar chart: solve time SGM by policy and difficulty."""
    if not HAS_MPL:
        return
    policies = [p for p in POLICY_ORDER if p in results]
    diffs    = [d for d in DIFFICULTIES if any(
                    d in results[p] for p in policies)]

    x     = np.arange(len(diffs))
    width = 0.8 / len(policies)
    fig, ax = plt.subplots(figsize=(10, 5))

    for i, policy in enumerate(policies):
        vals = []
        for diff in diffs:
            d = results[policy].get(diff, {})
            vals.append(sgm(d.get('times', [0])))
        offset = (i - len(policies)/2 + 0.5) * width
        bars = ax.bar(x + offset, vals, width, label=policy,
                      color=POLICY_COLORS.get(policy, 'gray'), alpha=0.85, edgecolor='white')
        for bar, val in zip(bars, vals):
            if not np.isnan(val):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                        f'{val:.1f}', ha='center', va='bottom', fontsize=8)

    ax.set_xlabel('Difficulty')
    ax.set_ylabel('Solve Time SGM (s)')
    ax.set_title(f'{BENCHMARK_NAMES.get(problem, problem)}: Solve Time Comparison')
    ax.set_xticks(x)
    ax.set_xticklabels([d.capitalize() for d in diffs])
    ax.legend(loc='upper left', fontsize=9)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()

    path = os.path.join(outdir, f'{problem}_solve_time.png')
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path}")


def fig_node_count_comparison(results, problem, outdir):
    """Bar chart: B&B node count SGM by policy and difficulty."""
    if not HAS_MPL:
        return
    policies = [p for p in POLICY_ORDER if p in results]
    diffs    = [d for d in DIFFICULTIES if any(d in results[p] for p in policies)]

    x     = np.arange(len(diffs))
    width = 0.8 / len(policies)
    fig, ax = plt.subplots(figsize=(10, 5))

    for i, policy in enumerate(policies):
        vals = []
        for diff in diffs:
            d = results[policy].get(diff, {})
            nodes = [n for n in d.get('nodes', []) if n is not None]
            vals.append(sgm(nodes) if nodes else float('nan'))
        offset = (i - len(policies)/2 + 0.5) * width
        bars = ax.bar(x + offset, vals, width, label=policy,
                      color=POLICY_COLORS.get(policy, 'gray'), alpha=0.85, edgecolor='white')

    ax.set_xlabel('Difficulty')
    ax.set_ylabel('B&B Nodes SGM')
    ax.set_title(f'{BENCHMARK_NAMES.get(problem, problem)}: B&B Node Count')
    ax.set_xticks(x)
    ax.set_xticklabels([d.capitalize() for d in diffs])
    ax.legend(loc='upper left', fontsize=9)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()

    path = os.path.join(outdir, f'{problem}_node_count.png')
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path}")


def fig_win_rate_heatmap(all_results, outdir):
    """Heatmap: Neural UCT win rate vs GCN+hybridestim across benchmarks × difficulties."""
    if not HAS_MPL:
        return

    problems = [p for p in BENCHMARKS if p in all_results]
    n_p = len(problems)
    n_d = len(DIFFICULTIES)
    data = np.full((n_p, n_d), np.nan)

    for i, prob in enumerate(problems):
        res = all_results[prob]
        baseline_key = "GCN + hybridestim"
        proposed_key = "GCN + Neural UCT"
        if baseline_key not in res or proposed_key not in res:
            continue
        for j, diff in enumerate(DIFFICULTIES):
            bt = res[baseline_key].get(diff, {}).get('times', [])
            pt = res[proposed_key].get(diff, {}).get('times', [])
            if bt and pt:
                data[i, j] = win_rate(bt, pt)

    fig, ax = plt.subplots(figsize=(7, max(3, n_p * 0.9)))
    im = ax.imshow(data, cmap='RdYlGn', vmin=0, vmax=1, aspect='auto')
    plt.colorbar(im, ax=ax, label='Win Rate (Neural UCT vs GCN+hybridestim)')

    ax.set_xticks(range(n_d))
    ax.set_xticklabels([d.capitalize() for d in DIFFICULTIES])
    ax.set_yticks(range(n_p))
    ax.set_yticklabels([BENCHMARK_NAMES.get(p, p) for p in problems])

    for i in range(n_p):
        for j in range(n_d):
            if not np.isnan(data[i, j]):
                color = 'white' if data[i, j] < 0.3 or data[i, j] > 0.7 else 'black'
                ax.text(j, i, f'{data[i,j]:.0%}', ha='center', va='center',
                        color=color, fontsize=11, fontweight='bold')

    ax.axvline(x=-0.5, color='white', linewidth=2)
    ax.set_title('Neural UCT Win Rate vs GCN + hybridestim\n(green = Neural UCT wins more)')
    plt.tight_layout()
    path = os.path.join(outdir, 'win_rate_heatmap.png')
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path}")


def fig_cumulative_solved(results, problem, difficulty, outdir):
    """Cumulative instances solved over time — like Figure 3 in the paper."""
    if not HAS_MPL:
        return
    policies = [p for p in POLICY_ORDER if p in results]
    fig, ax  = plt.subplots(figsize=(8, 5))

    time_limit = 3600
    time_bins  = np.linspace(0, time_limit, 200)

    for policy in policies:
        times = results[policy].get(difficulty, {}).get('times', [])
        if not times:
            continue
        cum = [(time_bins <= t).sum() / len(times) for t in times]
        sorted_times = np.sort(times)
        cum_solved   = np.arange(1, len(sorted_times) + 1) / len(sorted_times)
        ax.step(sorted_times, cum_solved, label=policy,
                color=POLICY_COLORS.get(policy, 'gray'), linewidth=2)

    ax.set_xlabel('Solve Time (s)')
    ax.set_ylabel('Fraction of Instances Solved')
    ax.set_title(f'{BENCHMARK_NAMES.get(problem, problem)} ({difficulty.capitalize()}): Cumulative Solved')
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)
    ax.set_xlim(0, time_limit)
    ax.set_ylim(0, 1.05)
    plt.tight_layout()
    path = os.path.join(outdir, f'{problem}_{difficulty}_cumulative.png')
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {path}")


# ── Excel workbook ─────────────────────────────────────────────────────────────

def save_excel(all_results, outpath):
    if not HAS_XLSX:
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(color="FFFFFF", bold=True)
    subhdr_fill  = PatternFill("solid", fgColor="2E75B6")
    subhdr_font  = Font(color="FFFFFF", bold=True)
    best_fill    = PatternFill("solid", fgColor="C6EFCE")
    border_thin  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    def style_cell(cell, fill=None, font=None, align='center', bold=False):
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        elif bold:
            cell.font = Font(bold=True)
        cell.border = border_thin

    # ── Sheet 1: Summary table per benchmark ──────────────────────────────────
    ws = wb.create_sheet("Summary")
    row = 1
    ws.cell(row, 1, "RankingBasedBB — Experimental Results Summary").font = Font(bold=True, size=14)
    ws.cell(row, 1).alignment = Alignment(horizontal='left')
    row += 2

    for problem, results in all_results.items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        c = ws.cell(row, 1, BENCHMARK_NAMES.get(problem, problem))
        style_cell(c, fill=header_fill, font=header_font)
        row += 1

        headers = ["Policy", "Easy Time SGM", "Med Time SGM", "Hard Time SGM",
                   "Easy Nodes", "Med Nodes", "Hard Nodes",
                   "Easy Solved", "Med Solved", "Hard Solved"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            style_cell(c, fill=subhdr_fill, font=subhdr_font)
        row += 1

        policies = [p for p in POLICY_ORDER if p in results]
        time_vals = {p: {} for p in policies}
        node_vals = {p: {} for p in policies}
        for p in policies:
            for diff in DIFFICULTIES:
                d = results[p].get(diff, {})
                time_vals[p][diff] = sgm(d.get('times', []))
                nodes = [n for n in d.get('nodes', []) if n is not None]
                node_vals[p][diff] = sgm(nodes) if nodes else float('nan')

        # Find best per column for highlighting
        best_time  = {d: min(time_vals[p][d] for p in policies if not np.isnan(time_vals[p].get(d, np.nan))) for d in DIFFICULTIES}
        best_nodes = {d: min(node_vals[p][d] for p in policies if not np.isnan(node_vals[p].get(d, np.nan))) for d in DIFFICULTIES}

        for p in policies:
            cells_data = [p]
            for diff in DIFFICULTIES:
                t = time_vals[p].get(diff, float('nan'))
                cells_data.append(f"{t:.2f}" if not np.isnan(t) else "N/A")
            for diff in DIFFICULTIES:
                n = node_vals[p].get(diff, float('nan'))
                cells_data.append(f"{n:.0f}" if not np.isnan(n) else "N/A")
            for diff in DIFFICULTIES:
                d = results[p].get(diff, {})
                cells_data.append(f"{d.get('solved',0)}/{d.get('total',0)}")

            for j, val in enumerate(cells_data, 1):
                c = ws.cell(row, j, val)
                is_best = False
                if j in [2,3,4]:   # time cols
                    diff = DIFFICULTIES[j-2]
                    try:
                        if float(val) <= best_time[diff] + 1e-6:
                            is_best = True
                    except Exception:
                        pass
                if j in [5,6,7]:   # node cols
                    diff = DIFFICULTIES[j-5]
                    try:
                        if float(val) <= best_nodes[diff] + 0.5:
                            is_best = True
                    except Exception:
                        pass
                style_cell(c, fill=best_fill if is_best else None,
                           align='center' if j > 1 else 'left')
            row += 1
        row += 1

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 16
    ws.column_dimensions['A'].width = 24

    # ── Sheet 2: GO/NO GO statistical tests ───────────────────────────────────
    ws2 = wb.create_sheet("GO_NO_GO")
    row = 1
    ws2.cell(row, 1, "GO/NO GO Statistical Tests (Wilcoxon + Win Rate)").font = Font(bold=True, size=14)
    row += 2

    hdrs = ["Benchmark", "Difficulty", "Comparison", "Win Rate", "p-value", "GO?", "Improvement %"]
    for j, h in enumerate(hdrs, 1):
        c = ws2.cell(row, j, h)
        style_cell(c, fill=subhdr_fill, font=subhdr_font)
    row += 1

    comparisons = [
        ("GCN + Neural UCT", "GCN + hybridestim", "Neural UCT vs GCN+hybridestim"),
        ("GCN + Neural UCT", "GCN + SCIP UCT",    "Neural UCT vs SCIP UCT"),
        ("GCN + Neural UCT", "SCIP Default",       "Neural UCT vs SCIP Default"),
    ]

    for problem, results in all_results.items():
        for diff in DIFFICULTIES:
            for proposed_key, baseline_key, label in comparisons:
                if proposed_key not in results or baseline_key not in results:
                    continue
                bt = results[baseline_key].get(diff, {}).get('times', [])
                pt = results[proposed_key].get(diff, {}).get('times', [])
                if not bt or not pt:
                    continue
                wr  = win_rate(bt, pt)
                pv  = wilcoxon_test(bt, pt)
                go  = wr >= 0.60 and pv < 0.05
                imp = improvement_pct(sgm(bt), sgm(pt))

                row_data = [
                    BENCHMARK_NAMES.get(problem, problem),
                    diff.capitalize(), label,
                    f"{wr:.1%}", f"{pv:.4f}",
                    "✓ GO" if go else "✗ NO GO",
                    f"{imp:.1f}%"
                ]
                for j, val in enumerate(row_data, 1):
                    c = ws2.cell(row, j, val)
                    fill = None
                    if j == 6:
                        fill = PatternFill("solid", fgColor="C6EFCE") if go else PatternFill("solid", fgColor="FFCCCC")
                    style_cell(c, fill=fill, align='center' if j > 1 else 'left')
                row += 1

    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 22

    # ── Sheet 3: Raw times ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Data")
    ws3.cell(1, 1, "Raw solve times per instance (all policies, all difficulties)").font = Font(bold=True)
    row = 3
    for problem, results in all_results.items():
        ws3.cell(row, 1, BENCHMARK_NAMES.get(problem, problem)).font = Font(bold=True)
        row += 1
        for diff in DIFFICULTIES:
            ws3.cell(row, 1, diff.capitalize()).font = Font(italic=True)
            col = 2
            for policy in POLICY_ORDER:
                if policy not in results:
                    continue
                times = results[policy].get(diff, {}).get('times', [])
                ws3.cell(row, col, policy).font = Font(bold=True)
                for i, t in enumerate(times, 1):
                    ws3.cell(row + i, col, round(t, 3))
                col += 1
            row += max(len(results[p].get(diff, {}).get('times', [])) for p in POLICY_ORDER if p in results) + 2

    wb.save(outpath)
    print(f"  Saved: {outpath}")


# ── Findings summary ───────────────────────────────────────────────────────────

def write_summary(all_results, outpath):
    lines = []
    lines.append("=" * 70)
    lines.append("  RankingBasedBB — Experimental Findings Summary")
    lines.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("=" * 70)
    lines.append("")
    lines.append("RESEARCH QUESTION")
    lines.append("  Does learned node selection (Neural UCT) improve B&B solve time")
    lines.append("  when variable selection is already optimized via Branch Ranking?")
    lines.append("")

    overall_go = []

    for problem, results in all_results.items():
        lines.append("-" * 70)
        lines.append(f"  {BENCHMARK_NAMES.get(problem, problem).upper()}")
        lines.append("-" * 70)

        baseline_key = "GCN + hybridestim"
        proposed_key = "GCN + Neural UCT"

        if baseline_key not in results or proposed_key not in results:
            lines.append("  [No results available]")
            continue

        for diff in DIFFICULTIES:
            bt = results[baseline_key].get(diff, {}).get('times', [])
            pt = results[proposed_key].get(diff, {}).get('times', [])
            if not bt or not pt:
                continue

            b_sgm = sgm(bt)
            p_sgm = sgm(pt)
            wr    = win_rate(bt, pt)
            pv    = wilcoxon_test(bt, pt)
            go    = wr >= 0.60 and pv < 0.05
            imp   = improvement_pct(b_sgm, p_sgm)
            overall_go.append(go)

            lines.append(f"\n  {diff.capitalize()} instances:")
            lines.append(f"    GCN + hybridestim  : {b_sgm:.2f}s SGM")
            lines.append(f"    GCN + Neural UCT   : {p_sgm:.2f}s SGM")
            lines.append(f"    Improvement        : {imp:.1f}%")
            lines.append(f"    Win rate           : {wr:.1%}  (threshold: 60%)")
            lines.append(f"    Wilcoxon p-value   : {pv:.4f}  (threshold: 0.05)")
            lines.append(f"    Decision           : {'✓ GO' if go else '✗ NO GO'}")

    lines.append("")
    lines.append("=" * 70)
    lines.append("  OVERALL CONCLUSION")
    lines.append("=" * 70)

    if not overall_go:
        lines.append("  No results available yet.")
    else:
        n_go  = sum(overall_go)
        n_tot = len(overall_go)
        lines.append(f"  GO decisions: {n_go}/{n_tot} ({n_go/n_tot:.0%})")
        if n_go / n_tot >= 0.6:
            lines.append("")
            lines.append("  RESULT: GO — Neural UCT provides statistically significant")
            lines.append("  improvement over SCIP's hybridestim when variable selection")
            lines.append("  is optimized. Learned node selection adds value beyond")
            lines.append("  optimal variable selection alone.")
        else:
            lines.append("")
            lines.append("  RESULT: NO GO — Neural UCT does not consistently outperform")
            lines.append("  SCIP's hybridestim. This validates the theoretical argument")
            lines.append("  that node selection is secondary to variable selection — even")
            lines.append("  with access to structural GCN embeddings, the improvement is")
            lines.append("  not statistically significant once branching is optimized.")
            lines.append("")
            lines.append("  This is a publishable negative result: first systematic")
            lines.append("  empirical validation of the implicit ML4CO assumption that")
            lines.append("  variable selection is the dominant bottleneck in B&B.")

    lines.append("")
    lines.append("  Files generated by this script:")
    lines.append("    results/figures/    — PNG figures")
    lines.append("    results/results.xlsx — Full data workbook")
    lines.append("    results/findings.txt — This summary")

    text = "\n".join(lines)
    with open(outpath, 'w') as f:
        f.write(text)
    print(f"  Saved: {outpath}")
    print()
    print(text)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--manual",  action="store_true", help="Use MANUAL_RESULTS dict")
    parser.add_argument("--problem", default=None, help="Single problem to analyse")
    parser.add_argument("--all",     action="store_true", help="All four benchmarks")
    parser.add_argument("--results-dir", default="results", help="Dir with .pkl result files")
    args = parser.parse_args()

    os.makedirs("results/figures", exist_ok=True)

    # ── Load data ──────────────────────────────────────────────────────────────
    if args.manual:
        all_results = MANUAL_RESULTS
        print("Using MANUAL_RESULTS — fill in the dict at the top of this file.")
    else:
        all_results = load_results(args.results_dir)
        if not all_results:
            print("No result pickle files found in results/. Using MANUAL_RESULTS template.")
            all_results = MANUAL_RESULTS

    if args.problem:
        all_results = {args.problem: all_results.get(args.problem, {})}

    if not any(all_results.values()):
        print("All result dicts are empty. Fill in MANUAL_RESULTS or save result pickles.")
        return

    print("\nGenerating figures...")
    for problem, results in all_results.items():
        if not results:
            continue
        fig_solve_time_comparison(results, problem, "results/figures")
        fig_node_count_comparison(results, problem, "results/figures")
        for diff in DIFFICULTIES:
            if any(diff in results.get(p, {}) for p in POLICY_ORDER):
                fig_cumulative_solved(results, problem, diff, "results/figures")

    fig_win_rate_heatmap(all_results, "results/figures")

    print("\nGenerating Excel workbook...")
    save_excel(all_results, "results/results.xlsx")

    print("\nGenerating findings summary...")
    write_summary(all_results, "results/findings.txt")

    print("\nDone. All outputs in results/")


if __name__ == "__main__":
    main()
